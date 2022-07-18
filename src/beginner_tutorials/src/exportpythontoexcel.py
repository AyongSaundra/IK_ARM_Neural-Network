#!/usr/bin/env python3

import numpy as np
import math
import PyKDL as kdl
from sklearn.preprocessing import Normalizer
import pickle
import rospy
from std_msgs.msg import Header, Float64
# from beginner_tutorials.msg import NN
from visualization_msgs.msg import Marker
# from geometry_msgs.msg import Point
import time
import openpyxl

# a1      = 0.1
# a2      = 0.16
# a3      = 0.089
# a4      = 0.105
# a5      = 0.021000
# a6      = 0.1736746

basetoj1    = 0.1
j1toj2      = 0.1
j2toj3      = 0.16
j3toj4      = 0.089
j4toj5      = 0.105
j5toj6      = 0.021000
j6toeff     = 0.1736746
links = [basetoj1,j1toj2,j2toj3,j3toj4,j4toj5,j5toj6,j6toeff]

#m
# a1      = 0.1
# a2      = 0.1
# a3      = 0.16
# a4      = 0.089
# a5      = 0.105
# a6      = 0.021000
# eff     = 0,1736746
# links = [a1,a2,a3,a4,a5,a6]

aaa = 1
Done = False
solved = 0
error = 0

def RX(yaw):
    return np.array([[1, 0, 0], 
                     [0, math.cos(yaw), -math.sin(yaw)], 
                     [0, math.sin(yaw), math.cos(yaw)]])   

def RY(delta):
    return np.array([[math.cos(delta), 0, math.sin(delta)], 
                     [0, 1, 0], 
                     [-math.sin(delta), 0, math.cos(delta)]])

def RZ(theta):
    return np.array([[math.cos(theta), -math.sin(theta), 0], 
                     [math.sin(theta), math.cos(theta), 0], 
                     [0, 0, 1]])

def TF(rot_axis=None, q=0, dx=0, dy=0, dz=0):
    if rot_axis == 'x':
        R = RX(q)
    elif rot_axis == 'y':
        R = RY(q)
    elif rot_axis == 'z':
        R = RZ(q)
    elif rot_axis == None:
        R = np.array([[1, 0, 0],
                      [0, 1, 0],
                      [0, 0, 1]])
    
    T = np.array([[R[0,0], R[0,1], R[0,2], dx],
                  [R[1,0], R[1,1], R[1,2], dy],
                  [R[2,0], R[2,1], R[2,2], dz],
                  [0, 0, 0, 1]])
    return T

def read_save_data():
    N = pickle.load(open("/home/ayong/arm_robot/arm_ws/src/beginner_tutorials/src/N.sav", 'rb'))
    regr = pickle.load(open("/home/ayong/arm_robot/arm_ws/src/beginner_tutorials/src/model.sav", 'rb'))
    X_train_norm = pickle.load(open("/home/ayong/arm_robot/arm_ws/src/beginner_tutorials/src/X_train_norm.sav", 'rb'))
    y_train = pickle.load(open("/home/ayong/arm_robot/arm_ws/src/beginner_tutorials/src/y_train.sav", 'rb'))
    # score = regr.score(X_train_norm, y_train)
    # print("Accuracy :", score)
    return N, regr, X_train_norm, y_train

def cekError(f_target, end_effector):
    f_result = end_effector
    f_diff = f_target.Inverse() * f_result
    [dx, dy, dz] = f_diff.p
    [drz, dry, drx] = f_diff.M.GetEulerZYX()
    error = np.sqrt(dx**2 + dy**2 + dz**2 + drx**2 + dry**2 + drz**2)
    error_pos = np.sqrt(dx**2 + dy**2 + dz**2)
    error_rot = np.sqrt(drx**2 + dry**2 + drz**2)
    error_list = [dx, dy, dz, drx, dry, drz]
    return error, error_list, error_pos, error_rot

def forward_kinematics(thetas):
    base = TF()
    joint1_from_base = TF(rot_axis='z', q=(thetas[0]), dz=(links[0]))
    joint1 = base.dot(joint1_from_base)
    joint2_from_joint1 = TF(rot_axis='y', q=(thetas[1]), dz=(links[1]))
    joint2 = joint1.dot(joint2_from_joint1)
    joint3_from_joint2 = TF(rot_axis='y', q=(thetas[2]), dz=(links[2]))
    joint3 = joint2.dot(joint3_from_joint2)
    joint4_from_joint3 = TF(rot_axis='z', q=(thetas[3]), dz=(links[3]))
    joint4 = joint3.dot(joint4_from_joint3)
    joint5_from_joint4 = TF(rot_axis='y', q=(thetas[4]), dz=(links[4]))
    joint5 = joint4.dot(joint5_from_joint4)
    joint6_from_joint5 = TF(rot_axis='z', q=(thetas[5]), dz=(links[5]))
    joint6 = joint5.dot(joint6_from_joint5)
    end_effector_from_joint6 = TF(dz=links[6])
    end_effector = joint6.dot(end_effector_from_joint6)

    f_eff = kdl.Frame(kdl.Rotation(end_effector[0,0], end_effector[0,1], end_effector[0,2],
                                   end_effector[1,0], end_effector[1,1], end_effector[1,2],
                                   end_effector[2,0], end_effector[2,1], end_effector[2,2]),
                        kdl.Vector(end_effector[0,3], end_effector[1,3], end_effector[2,3]))

    x, y, z = f_eff.p
    # roll, pitch, yaw = f_eff.M.GetRPY()
    qx, qy, qz, qw = f_eff.M.GetQuaternion()
    return x, y, z, qx, qy, qz, qw, f_eff

def init_publisher():
    # global rate
    # rospy.init_node('joint_positions_node', anonymous=True)
    # rate = rospy.Rate(1)
    pub1 = rospy.Publisher('/arm_robot/link1_joint_position_controller/command', Float64, queue_size=100)
    pub2 = rospy.Publisher('/arm_robot/link2_joint_position_controller/command', Float64, queue_size=100)
    pub3 = rospy.Publisher('/arm_robot/link3_joint_position_controller/command', Float64, queue_size=100)
    pub4 = rospy.Publisher('/arm_robot/link4_joint_position_controller/command', Float64, queue_size=100)
    pub5 = rospy.Publisher('/arm_robot/link5_joint_position_controller/command', Float64, queue_size=100)
    pub6 = rospy.Publisher('/arm_robot/link6_joint_position_controller/command', Float64, queue_size=100)
    return pub1, pub2, pub3, pub4, pub5, pub6

def publish_joint(pub1, pub2, pub3, pub4, pub5, pub6, joints):
    # rospy.loginfo(joints[0,0])
    # rospy.loginfo(joints[0,1])
    # rospy.loginfo(joints[0,2])
    # rospy.loginfo(joints[0,3])
    # rospy.loginfo(joints[0,4])
    # rospy.loginfo(joints[0,5])
    # print("============================================")
    # print("Target frame :")
    # print(end_effector)
    # print("============================================")
    for i in range(8):
        pub1.publish(joints[0])
        pub2.publish(joints[1])
        pub3.publish(joints[2])
        pub4.publish(joints[3])
        pub5.publish(joints[4])
        pub6.publish(joints[5])
        # rate.sleep()

def publish_marker(pose,color):
    global aaa,solved, error, stat
    marker = Marker()
    marker.header.frame_id = "base"
    marker.type = Marker.CUBE
    marker.action = Marker.ADD
    marker.id = aaa
    marker.scale.x = 0.035
    marker.scale.y = 0.035
    marker.scale.z = 0.035
    if color == 0:
        marker.color.r = 0
        marker.color.g = 1
        marker.color.b = 0
        solved += 1
        stat = "berhasil"
    elif color == 1:
        marker.color.r = 1
        marker.color.g = 0
        marker.color.b = 0
        error +=1
        stat = "gagal"

    marker.color.a = 1.0
    
    marker.pose.position.x = pose[0]
    marker.pose.position.y = pose[1]
    marker.pose.position.z = pose[2]

    marker.pose.orientation.x = pose[3]
    marker.pose.orientation.y = pose[4]
    marker.pose.orientation.z = pose[5]
    marker.pose.orientation.w = pose[6]

    pub = rospy.Publisher('/visualization_marker', Marker, queue_size=1)
    pub.publish(marker)
    aaa+=1

def subtitleexcel(my_sheet,n_marker ,n_data, neuronperlayer, score, mse, rmse, mae):
    c1 = my_sheet['B1']
    c1.value = "Total Train Data :"
    c2 = my_sheet['C1']
    c2.value = n_data

    c3 = my_sheet['B2']
    c3.value = "Neuron per-Layer :"
    c4 = my_sheet['C2']
    c4.value = neuronperlayer

    c5 = my_sheet['B3']
    c5.value = "Total Marker :"
    c6 = my_sheet['C3']
    c6.value = n_marker

    c7 = my_sheet['B4']
    c7.value = "Err Threshold(cm/rad) :"
    c8 = my_sheet['C4']
    c8.value = "5/0.1"

    c9 = my_sheet['D1']
    c9.value = "Score :"
    c10 = my_sheet['E1']
    c10.value = score
    c11 = my_sheet['D2']
    c11.value = "MSE :"
    c12 = my_sheet['E2']
    c12.value = mse
    c13 = my_sheet['D3']
    c13.value = "RMSE :"
    c14 = my_sheet['E3']
    c14.value = rmse
    c15 = my_sheet['D4']
    c15.value = "MAE :"
    c16 = my_sheet['E4']
    c16.value = mae

    c17 = my_sheet['A7']
    c17.value = "Iter ke"
    c18 = my_sheet['B7']
    c18.value = "Error Position(cm)"
    c19 = my_sheet['C7']
    c19.value = "Error Rotation(rad)"
    c20 = my_sheet['D7']
    c20.value = "Status"
    c21 = my_sheet['E7']
    c21.value = "Time per-Marker(s)"

def main():
    rospy.init_node('marker_basic_node', anonymous=True)
    rate = rospy.Rate(0.5)
    pub1, pub2, pub3, pub4, pub5, pub6 = init_publisher()
    N, regr, _, _ = read_save_data()

    n_marker = 100
    errpos_threshold = 1
    errrot_threshold = 0.05
    n_data = 50000
    neuronperlayer = "4096,4096,4096,4096,4096"
    score = 0.95
    mse = 0.085
    rmse = 0.290
    mae = 0.225

    iter_ke = 1
    start_row = 8
    erroravg = 0
    total_waktu = []

    my_wb = openpyxl.Workbook()
    my_sheet = my_wb.active
    subtitleexcel(my_sheet,n_marker ,n_data, neuronperlayer, score, mse, rmse, mae)

    for i in range(n_marker):
        #input sudut tiap joint secara random
        thetas = [np.random.uniform(-np.pi,np.pi),
                  np.random.uniform(-np.pi,np.pi),
                  np.random.uniform(-np.pi,np.pi),
                  np.random.uniform(-np.pi,np.pi),
                  np.random.uniform(-np.pi,np.pi),
                  np.random.uniform(-np.pi,np.pi)]
        delta_q = np.random.normal(loc=[0, 0, 0, 0, 0, 0], scale=[0.33, 0.33, 0.33, 0.33, 0.33, 0.33], size=None)
        q_current = np.clip((thetas + delta_q), a_min=[-np.pi, -np.pi, -np.pi, -np.pi, -np.pi, -np.pi], a_max=[np.pi, np.pi, np.pi, np.pi, np.pi, np.pi])

        xi, yi, zi, qxi, qyi, qzi, qwi, _ = forward_kinematics(thetas)
        xc, yc, zc, qxc, qyc, qzc, qwc, f_effc = forward_kinematics(q_current)

        pose_init = np.array([xi, yi, zi, qxi, qyi, qzi, qwi])
        pose_current = np.array([xc, yc, zc, qxc, qyc, qzc, qwc])
        delta_pose = pose_current - pose_init

        X = np.hstack((thetas, delta_pose))
        input = X.reshape(1,13)
        
        start_time = time.time()
        joints = regr.predict(input)
        waktu=(" %s " % round(time.time() - start_time, 4))
        print(f"{waktu} seconds")
        waktu = float(waktu)

        joints = [joints[0,0]+thetas[0],joints[0,1]+thetas[1],joints[0,2]+thetas[2],joints[0,3]+thetas[3],joints[0,4]+thetas[4],joints[0,5]+thetas[5]]
        pvx, pvy, pvz, pqx, pqy, pqz, pqw, f_effj = forward_kinematics(joints)
        pose = np.array([pvx, pvy, pvz, pqx, pqy, pqz, pqw])
        # print("Random Position :", X)
        # print("Predict Position :", pose)
        # print("Joints in deg:", joints)

        _,_,c,d = cekError(f_effc,f_effj)
        # print("error avg :", a)
        # print("error list :", b)
        # print("error pos m :", c)
        
        c = c*100 #convertmtocm

        publish_joint(pub1, pub2, pub3, pub4, pub5, pub6, joints)
        if c <= errpos_threshold or d <= errrot_threshold:
            publish_marker(pose,0)
        else:
            publish_marker(pose,1)
        solvability_rate = solved/n_marker
        solvability_rate = solvability_rate*100
        print("Total marker : ", n_marker)
        print("iterasi ke : ", iter_ke)
        print("Berhasil", solved)
        print("Gagal", error)
        print("error pos :", c, "cm")
        print("error rot :", d, "rad")
        print("solvability_rate : %.0f"% solvability_rate)
        print("====================================")

        w = my_sheet.cell(row= start_row , column = 1)
        w.value = iter_ke
        w1 = my_sheet.cell(row= start_row , column = 2)
        w1.value = c
        w2 = my_sheet.cell(row= start_row , column = 3)
        w2.value = d
        w3 = my_sheet.cell(row= start_row , column = 4)
        w3.value = stat
        w3 = my_sheet.cell(row= start_row , column = 5)
        w3.value = waktu

        total_waktu.append(waktu)
        iter_ke += 1
        start_row += 1
        rate.sleep()
    
    # total_waktu = sum(waktu)

    w4 = my_sheet.cell(row= start_row+2 , column = 1)
    w4.value = "Total Berhasil :"
    w5 = my_sheet.cell(row= start_row+2 , column = 2)
    w5.value = solved
    w6 = my_sheet.cell(row= start_row+3 , column = 1)
    w6.value = "Solvability Rate(%)"
    w7 = my_sheet.cell(row= start_row+3 , column = 2)
    w7.value = solvability_rate
    w8 = my_sheet.cell(row= start_row+4 , column = 1)
    w8.value = "Marker time(ms)"
    # w9 = my_sheet.cell(row= start_row+4 , column = 2)
    # w9.value = total_waktu

    my_wb.save("IKNNdata1_005.xlsx")

if __name__ == "__main__":
    main()